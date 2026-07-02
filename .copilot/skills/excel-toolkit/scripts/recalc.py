#!/usr/bin/env python3
"""
Excel Formula Recalculation Script
Recalculates all formulas in an Excel file using LibreOffice.
Usage: python recalc.py <excel_file> [timeout_seconds]
"""
import json, sys, subprocess, os, platform
from pathlib import Path
from openpyxl import load_workbook


def setup_libreoffice_macro():
    if platform.system() == 'Darwin':
        macro_dir = os.path.expanduser('~/Library/Application Support/LibreOffice/4/user/basic/Standard')
    else:
        macro_dir = os.path.expanduser('~/.config/libreoffice/4/user/basic/Standard')
    
    macro_file = os.path.join(macro_dir, 'Module1.xba')
    if os.path.exists(macro_file):
        with open(macro_file, 'r') as f:
            if 'RecalculateAndSave' in f.read():
                return True
    
    if not os.path.exists(macro_dir):
        subprocess.run(['soffice', '--headless', '--terminate_after_init'], capture_output=True, timeout=10)
        os.makedirs(macro_dir, exist_ok=True)
    
    macro_content = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>'''
    
    try:
        with open(macro_file, 'w') as f:
            f.write(macro_content)
        return True
    except Exception:
        return False


def recalc(filename, timeout=30):
    if not Path(filename).exists():
        return {'error': f'File {filename} does not exist'}
    
    abs_path = str(Path(filename).absolute())
    if not setup_libreoffice_macro():
        return {'error': 'Failed to setup LibreOffice macro'}
    
    cmd = [
        'soffice', '--headless', '--norestore',
        'vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application',
        abs_path
    ]
    
    if platform.system() != 'Windows':
        timeout_cmd = 'timeout' if platform.system() == 'Linux' else None
        if platform.system() == 'Darwin':
            try:
                subprocess.run(['gtimeout', '--version'], capture_output=True, timeout=1, check=False)
                timeout_cmd = 'gtimeout'
            except (FileNotFoundError, subprocess.TimeoutExpired):
                pass
        if timeout_cmd:
            cmd = [timeout_cmd, str(timeout)] + cmd
    
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode not in (0, 124):
        return {'error': result.stderr or 'Unknown error during recalculation'}
    
    try:
        wb = load_workbook(filename, data_only=True)
        excel_errors = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']
        error_details = {err: [] for err in excel_errors}
        total_errors = 0
        
        for sheet_name in wb.sheetnames:
            for row in wb[sheet_name].iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                error_details[err].append(f"{sheet_name}!{cell.coordinate}")
                                total_errors += 1
                                break
        wb.close()
        
        res = {'status': 'success' if total_errors == 0 else 'errors_found', 'total_errors': total_errors, 'error_summary': {}}
        for err_type, locations in error_details.items():
            if locations:
                res['error_summary'][err_type] = {'count': len(locations), 'locations': locations[:20]}
        
        wb2 = load_workbook(filename, data_only=False)
        formula_count = sum(
            1 for sn in wb2.sheetnames for row in wb2[sn].iter_rows()
            for cell in row if cell.value and isinstance(cell.value, str) and cell.value.startswith('=')
        )
        wb2.close()
        res['total_formulas'] = formula_count
        return res
    except Exception as e:
        return {'error': str(e)}


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        sys.exit(1)
    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    print(json.dumps(recalc(filename, timeout), indent=2))

if __name__ == '__main__':
    main()
