#!/usr/bin/env python3
"""
Inspect an Excel file and return structured metadata as JSON.
Usage: python inspect_excel.py <file.xlsx> [--data] [--sheet SheetName] [--rows N]
  --data    Include cell data preview (first N rows, default 20)
  --sheet   Inspect specific sheet only
  --rows N  Number of preview rows (default 20)
"""
import json, sys, argparse
from pathlib import Path

def inspect(filepath, include_data=False, sheet_name=None, max_rows=20):
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=False)
    
    result = {
        "file": str(filepath),
        "sheets": [],
        "total_sheets": len(wb.sheetnames),
    }
    
    sheets_to_inspect = [sheet_name] if sheet_name else wb.sheetnames
    
    for name in sheets_to_inspect:
        if name not in wb.sheetnames:
            result["sheets"].append({"name": name, "error": "Sheet not found"})
            continue
        ws = wb[name]
        info = {
            "name": name,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
        }
        
        # Get headers (first row)
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(c) if c is not None else "" for c in row]
        info["headers"] = headers
        
        if include_data:
            rows = []
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows + 1, values_only=True)):
                rows.append([str(c) if c is not None else "" for c in row])
            info["data_preview"] = rows
        
        # Detect column types from first 100 rows
        if ws.max_row and ws.max_row > 1:
            col_types = {}
            for row in ws.iter_rows(min_row=2, max_row=min(101, ws.max_row + 1)):
                for cell in row:
                    col = cell.column
                    if cell.value is not None:
                        t = type(cell.value).__name__
                        col_types.setdefault(col, set()).add(t)
            info["column_types"] = {
                headers[k-1] if k-1 < len(headers) else f"Col{k}": list(v)
                for k, v in col_types.items()
            }
        
        result["sheets"].append(info)
    
    wb.close()
    return result

def main():
    parser = argparse.ArgumentParser(description="Inspect Excel file structure")
    parser.add_argument("file", help="Excel file path")
    parser.add_argument("--data", action="store_true", help="Include data preview")
    parser.add_argument("--sheet", help="Specific sheet name")
    parser.add_argument("--rows", type=int, default=20, help="Preview rows (default 20)")
    args = parser.parse_args()
    
    if not Path(args.file).exists():
        print(json.dumps({"error": f"File not found: {args.file}"}))
        sys.exit(1)
    
    result = inspect(args.file, args.data, args.sheet, args.rows)
    print(json.dumps(result, indent=2, default=str))

if __name__ == "__main__":
    main()
